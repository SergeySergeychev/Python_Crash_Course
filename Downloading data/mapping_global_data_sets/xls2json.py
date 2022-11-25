
import json
import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook


# Open the workbook and select a worksheet.
wb = load_workbook('gdp_by_country.xlsx')
sheet = wb['Sheet1']

# List to hold dictionaries.
country_list = []

# Iterate through each row in worksheet and fetch values into dict.
for row in islice(sheet.values, 1, sheet.max_row):
    country = OrderedDict()
    country['Country Name'] = row[1]
    country['Year'] = row[4]
    country['GDP(US)'] = row[5]
    country_list.append(country)

# Serialize the list of dicts to json
j = json.dumps(country_list)

# Write to file
with open('gdp_by_country_2.json', 'w') as f:
    f.write(j)
