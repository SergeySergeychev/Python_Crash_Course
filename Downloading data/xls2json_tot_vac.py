import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

# Open the workbook and select a worksheet.
wb = load_workbook('vaccinations.xlsx')
sheet = wb['vaccinations']

# List to hold dicitionaris

vaccination_list = []

# iterate through each row in worksheet and fetch values into dict.
for row in islice(sheet.values, 1, sheet.max_row):
    vacc = OrderedDict()
    vacc['Country Name'] = row[0]
    vacc['Date'] = row[2]
    vacc['Total Vaccinations'] = row[4]
    vaccination_list.append(vacc)

# Serialize the list of dicts to json.
j = json.dumps(vaccination_list, indent=4, sort_keys=True, default=str)

# Write to file.
with open('vaccination.json', 'w') as f:
    f.write(j)
